# coding: utf-8

import shutil
import openpyxl
import configparser
import zipfile,os
import datetime
import time
import sys
import os

''''
configsetting
''' 
config_ini = configparser.ConfigParser()
config_ini.read('config.ini', encoding='utf-8')
iwb = config_ini.get('DEFAULT', 'InputWorkbook')
isn = config_ini.get('DEFAULT', 'InputSheetName')


'''
excellsetting
'''
inputWb = openpyxl.load_workbook(iwb,keep_vba=True)
inputWs = inputWb[ isn ]

'''
エクセルの値取得
'''
fileNumberLimit      = inputWs.cell(row = 20, column = 3).value
pictureFilePath      = inputWs.cell(row =  9, column = 3).value
zipSaveFilePath      = inputWs.cell(row = 10, column = 3).value
folderLimtTotalSize  = inputWs.cell(row = 21, column = 3).value
executeInterval      = inputWs.cell(row = 18, column = 3).value
topZipFileName       = inputWs.cell(row = 16, column = 3).value

print(folderLimtTotalSize)
print(fileNumberLimit)
print(executeInterval)
print(pictureFilePath)
print(zipSaveFilePath)
print(topZipFileName)


"""
zipファイルの数を取得する関数
拡張子がzipのものをカウントアップする。
"""
def fileNumberReserch(zipSaveFilePath, readZipCounter):
    #ジップファイルの数を計算
    readZipCounter = 0
    for file in os.listdir(zipSaveFilePath):
        base, ext = os.path.splitext(file)
        if ext == '.zip':
            readZipCounter += 1
    print("ZIPファイル数"+ str(readZipCounter) + "個")   
    return  readZipCounter


"""
指定のフォルダの全.jpgを取得する
@ param  picturePath
@ rerutn jpgList

"""
def getJpgList(picturePath):   
    #jpgフォルダを指定してリストを取得する・
    jpgList = os.listdir(picturePath)
    print(jpgList) 

    #jpgリストから.jpg拡張子以外のものを削除する。
    for i in jpgList:
        print(i)
        if "jpg" not in i:
            jpgList.remove(i)
    return jpgList


"""
ファイルの容量を取得する関数
@param  file ファイル名
@rerutn size MB
"""
def getImageSize(pictureFilePath,file):
    path = pictureFilePath + "\\" +str(file)
    size = os.path.getsize(path) / 1024 / 1024
    return size


"""
main関数
"""
shutil.rmtree(zipSaveFilePath)
os.mkdir(zipSaveFilePath)
jpgList = getJpgList(pictureFilePath)
num = 1
dateName        = datetime.datetime.now()

while(True):
    #1～rangeまで出力
    print(str(num) +"回目の処理を開始します。")

    
    """
    時刻の取得
    """
    #時刻の取得
    
    dateName = dateName + datetime.timedelta(minutes = executeInterval)
    topDateAddName  = dateName.strftime('%Y%m%d%H%M%S')
    print(topDateAddName)
    
    """
    ZIPの数を取得
    """
    readCounter = 0
    readCounter =  fileNumberReserch(zipSaveFilePath, readCounter)
    
        #ファイルサイズ取得

    FileSize = 0
    counter  = 0
    i        = -1

    #指定の上限数よりカウンター小さいときに処理を継続する。
    print("指定のZIPファイル上限数と現状作成完了しているZip数を比較します")
    if(readCounter <= fileNumberLimit -1 ):
        
        #ファイル名を指定してZip化する。
        with zipfile.ZipFile(str(zipSaveFilePath) + '\\' + topZipFileName + topDateAddName + '.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:

            #ファイルパスを指定してループする
            while True:
                i = i + 1
                print(str(i) +"インデックス")
                try:
                    FileSize = getImageSize( pictureFilePath,jpgList[0]) + FileSize
                    counter = counter + 1
                    print(str(FileSize) +"Filesize")
                    print(str(counter)+"counter数")
                    os.chdir(pictureFilePath)

                    #サイズオーバーしたときにひとつ前までのリストを削除する。
                    if(FileSize < folderLimtTotalSize):
                        print("")
                        zipf.write( jpgList[0] , compress_type=zipfile.ZIP_DEFLATED)
                        jpgList.remove(jpgList[0])

                        jpgListNumber =  len(jpgList)


                    else:
                        #一件目でファイル容量オーバーの場合
                        if counter == 1:
                            print("指定のフォルダサイズよりも大きいファイルが存在します。")
                            sys.exit()
                        break
                except IndexError:
                    print("処理を終了します。")
                    sys.exit()
                    

        zipf.close()
    else:
        print("zipファイルが上限に達しました。")
        break

    #インクリメント
    print(str(num) +"回目の処理を終了します。")
    num = num + 1
    #    time.sleep(executeInterval * 60)
    