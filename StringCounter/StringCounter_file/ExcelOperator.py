import openpyxl


class ExcelOperator:
    
    """
    コンストラクタ
    file名とシート名を取得
    @ fileName = エクセルファイルパス
    @ fileName = エクセルシート名
    """
    def __init__(self,fileName ):  
        self.fileName  =  openpyxl.load_workbook( fileName )
        print(self.fileName)
        self.SheetName =  self.fileName.worksheets[ 0 ]
        print(self.SheetName)
    
    
    """
    loadValue
    gyouとretuから値を取得する。
    @ gyou 行番号
    @ retu 列番号
    """
    def loadValue(self, gyou , retu):
        URL_Value      = self.SheetName.cell(row = gyou, column = retu).value
        return URL_Value


    """
    getRow
    一番下の行の座標を取得する。
    @ return max 最終行
    """
    def getRow(self):
        max = self.SheetName.max_row
        print(str(max) + '行')
        return max


    """
    writeValue
    値を書き込みする
    @ return value ,gyou ,retu
    """
    def writeValue(self, value , gyou , retu):
        self.SheetName.cell(row = gyou, column = retu).value =  value
        self.fileName.save("Count.xlsx")