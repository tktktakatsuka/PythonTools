
import os
# resuests モジュールをインポート
import re
import time
from tkinter import messagebox
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import random
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
import requests
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles.borders import Border, Side
import datetime

"""
isStatusCodeNomal
@param URL 取得するURL
""" 
def isStatusCodeNomal(URL):
    res = requests.get(URL)
    isStatus =  res.status_code
    print(res.status_code)
    return  isStatus
  
  
"""
driverOpen
@param URL 取得するURL
""" 
def driverOpen(URL):
    options = Options()
    user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36']
    options = webdriver.ChromeOptions()
    options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
    res = driver.get(URL)
      
"""
getCurrentPage
@param className HTMLページのクラスを指定
@return tagList
"""   
def getCurrentTextPage(className):
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    tagList = soup.select(className)
    
    textlist = [] 
    for item in tagList:
        textlist.append( item.getText()) 
    return textlist 

def getCurrentURLPage(className):
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    tagList = soup.select(className)
    
    urlList= []
    for item in tagList:
        url =  item.find('a')
        url = url.get('href')
        urlList.append( url)
    
    return urlList

def getgigityukiPage():
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    tagList = soup.select('//*[@id="wrapper"]/main/article/section[1]/div[2]/div[14]/p/span[2]')
    
    for item in tagList:
        url =  item.find('a').text
    return url


'''
画面遷移
'''
def moveGamen(xpath):
    elem = driver.find_element(By.XPATH,xpath)
    elem.click()   

import sqlite3
from unicodedata import name

# TEST.dbを作成する
# すでに存在していれば、それにアスセスする。
def dbAccess(dbname):
    conn = sqlite3.connect(dbname) 
    return conn


def insert_data(str銘柄,str企業名,str優待内容 ,month1 ,month2 ,必要投資金額 ,優待利回り ,配当利回り,url):

    try:
        sql = f"INSERT INTO yutai (企業名, 銘柄コード,優待内容,権利確定月①,権利確定月②,必要投資金額,優待利回り,配当利回り,URL) values (?,?,?,?,?,?,?,?,?)"     
        data = [str(str企業名),str(str銘柄),str(str優待内容),str(month1),str(month2),str(必要投資金額),str(優待利回り),str(配当利回り),url]
        cur.execute(sql, data)
        con.commit()
    except sqlite3.IntegrityError:
        global uniqueCounter
        uniqueCounter = uniqueCounter +1 
        pass
    
    
"""
mainMethod
@param void
"""
if __name__ == '__main__':
    print("処理を開始します。")
    time.sleep(2)
    
    
    #グローバル変数
    with webdriver.Chrome(ChromeDriverManager().install()) as driver:
        i = 0
    
    
        con = sqlite3.connect('yuutai.sqlite3')
        cur = con.cursor()
        cur.execute('create table IF NOT EXISTS yutai(企業名 , 銘柄コード unique , 優待内容  , 権利確定月① , 権利確定月② , 必要投資金額  , 優待利回り , 配当利回り,URL )')
        uniqueCounter = 0
        
        MONTHList = ["january", 
                     "february" ,
                     "march" ,
                     "april" ,
                     "may" , 
                     "june" ,
                     "july" ,
                     "august" , 
                     "september" , 
                     "october" ,
                     "november" , 
                     "december"  ]

        for mList in MONTHList:
            
            while(True):
                time.sleep(0.4)
                if i == 0:
                    URL ="https://www.kabuyutai.com/yutai/"+ mList +".html"
                else:
                    URL ="https://www.kabuyutai.com/yutai/"+ mList + str(i + 1) +".html"

                #ページがなかったら処理終了
                if(isStatusCodeNomal(URL) == 404):
                    i=0
                    break
                
                #URLを開く
                driverOpen(URL)

                #指定のURLのクラスからデータを取得する
                className = "[class*='table_tr_info']"
                taglist = getCurrentTextPage(className)
                

                
                urllist = getCurrentURLPage(className)

                coutner = 0
                #書き込みしたい文字列を表示するところまで完了
                for item in taglist:
                    
                    
                    url = urllist[coutner]

                    #必要な情報を取得    
                    splitedList = str(item).splitlines()
                    
                    #infoがあった場合に削除するためのif分
                    if '長期優待を狙えば、' in splitedList[2]:
                        splitedList.pop(2) #これでインデックス番号2番の削除後詰める
                    #長期優待を削除するためのif分
                    for item in taglist:
                        
                        if '【長期優待】' in splitedList[2]:
                            splitedList.pop(2) #これでインデックス番号2番の削除後詰める
                        
                        if '【長期優待】' in splitedList[3]:
                            splitedList.pop(3) #これでインデックス番号2番の削除後詰める



                                    
                    pattern = r'\d+'
                    r銘柄 = re.search(pattern, splitedList[1])
                    #書き込むもの
                    str銘柄 = r銘柄.group()
                    print(str銘柄)

                    pattern = r'\D+'
                    r企業名 = re.search(pattern, splitedList[1])
                    #書き込むもの
                    str企業名 = r企業名.group().replace("（" , "")
                    print(str企業名)


                    pattern = r'(【.*?】)'
                    r優待内容 = re.search(pattern, splitedList[2])
                    #書き込むもの
                    str優待内容 =splitedList[2].replace(r優待内容.group(),"" )
                    #test
                    #str優待内容 ="test"
                    print(str優待内容)

                    month1=""
                    month2=""
                    pattern = r'(【.*?】)'
                    r確定月 = re.search(pattern, splitedList[3])
                    month  = splitedList[3].replace(r確定月.group(),"" ).replace("月","" )
                    monthList = month.split("・")
                    for num in range(len(monthList)):
                        if num == 0:
                            month1 = monthList[num]
                        if num == 1:
                            month2 = monthList[num]
                        #書き込むもの
                        print(monthList[num])

                    pattern = r'(【.*?】)'
                    r必要投資金額 = re.search(pattern, splitedList[4])
                    必要投資金額  = splitedList[4].replace(r必要投資金額.group(),"" ).replace("円","" )
                    必要投資金額list = 必要投資金額.split("・")
                    必要投資金額 = 必要投資金額list[0]
                    #書き込むもの
                    print(必要投資金額list[0])

                    pattern = r'(【.*?】)'
                    r優待利回り = re.search(pattern, splitedList[5])
                    優待利回り  = splitedList[5].replace(r優待利回り.group(),"" ).replace("","" )
                    優待利回りlist = 優待利回り.split("・")
                    優待利回り = 優待利回りlist[0]
                    #書き込むもの
                    print(優待利回りlist[0])

                    pattern = r'(【.*?】)'
                    r配当利回り = re.search(pattern, splitedList[6])
                    配当利回り  = splitedList[6].replace(r配当利回り.group(),"" ).replace("","" )
                    配当利回りlist = 配当利回り.split("・")
                    配当利回り = 配当利回りlist[0]
                    #書き込むもの
                    print(配当利回りlist[0])

                    insert_data(str銘柄,str企業名,str優待内容 ,month1 ,month2 ,必要投資金額 ,優待利回り ,配当利回り, url)
                    coutner= coutner +1
                i= i+1
                
    # 取り出したデータをExcelに貼り付ける --- (*4)


    dt_now = datetime.datetime.now()
    inputBookName = "Yutai_List.xlsm"
    book = openpyxl.load_workbook(inputBookName, keep_vba=True) # 新規ブック作成
    sheet = book.worksheets[0] # 先頭のシート
    sql = 'SELECT * FROM yutai'
    rows = cur.execute(sql)
    # 罫線(外枠)を設定
    border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))
    
    sheet.cell(1, 2).value = dt_now.strftime('%Y年%m月%d日') # 単語
    for i, n in enumerate(rows): # --- (*4a)
    
        sheet.cell(i+5, 1).value = n[0] # 単語
        sheet.cell(i+5, 2).value = n[1] # 意味
        sheet.cell(i+5, 3).value = n[2] # 意味
        sheet.cell(i+5, 4).value = n[3] # 意味
        sheet.cell(i+5, 5).value = n[4] # 意味
        sheet.cell(i+5, 6).value = n[5] # 意味
        sheet.cell(i+5, 7).value = n[6] # 意味
        sheet.cell(i+5, 8).value = n[7] # 意味
        sheet.cell(i+5, 9).hyperlink = n[8] # 意味
        
    # set column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
    
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
    
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    #最大行
    maxRow = sheet.max_row + 1
    
    # セルに罫線を設定
    for row_num in range(5,maxRow):
        for col_num in range(1,10):
            sheet.cell(row=row_num ,column=col_num).border = border
        
    # Excelファイルに保存 --- (*5)
    book.save(inputBookName)
    

messagebox.showinfo("正常", "処理を正常終了します。")
    
