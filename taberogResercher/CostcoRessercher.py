
import random
from webdriver_manager.chrome import ChromeDriverManager
from operator import truediv
from bs4 import BeautifulSoup
from bs4.builder import HTML
from openpyxl.xml.constants import ACTIVEX, XLSX
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from tkinter import *

"""
driversetting
@param URL 取得するURL
""" 
def driverSetting(URL):
    options = Options()
    user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36']
    options = webdriver.ChromeOptions()
    options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
    driver.get(URL)
    
"""
getCurrentPage
@param className HTMLページのクラスを指定
@return tagList
"""   
def getCurrentPage(className):
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    tagList = soup.select(className)
    return tagList

"""
getItem
@param tagList HTMLのリストの中の特定のクラスを抜き出す。
@return ItemList
"""   
def getItem(tagList , tag , classname):

    itemList = [a,b,c]
    for item in tagList:
        
        #アイテム変数定義
        info   = item.find(tag , class_=classname ).text
        info   = item.find(tag , class_=classname ).text
        info   = item.find(tag , class_=classname ).text
        
        
        return info

"""
mainMethod
@param void
"""
if __name__ == '__main__':
    #グローバル変数
    driver = webdriver.Chrome(ChromeDriverManager().install())
    
    #指定のURLからブラウザを開く
    URL ="https://tabelog.com/aichi/A2301/A230101/23074669/"
    driverSetting(URL)
    
    #指定のURLのクラスからデータを取得する
    className = "[class*='rstdtl-rstdata box-contents']"
    taglist = getCurrentPage(className)
    print(taglist)
    
    
    #指定のタグから情報を取得する。
    print("\n\n\n")
    print(getItem(taglist , 'div', 'rstinfo-table__name-wrap'))
    
    #ドライバー終了
    driver.quit()



"""
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    taglist = soup.select("[class*='order-history__list-item']")

    URL抽出

    for item in taglist:

        アイテム変数定義

        pathNumber   = item.find('span', class_='order-id notranslate' ).text
        price        = item.find('div', class_='list-item__data ordr-history-price' ).text
        status       = item.find('div', class_='list-item__data bold' ).text.replace(' ','').replace('\n','')
        path         = 'https://www.costco.co.jp/my-account/order/' + pathNumber.replace('\n','')
        
        #content = status
        #print(content)
        #pattern = '.*返金済み.*'
        #result = re.fullmatch(pattern, str(content))
            
        #価格        
        if not status == "返金済み":
            inputWs.cell(row = num, column = yoko +6 ).value =  str(price)
            pathList.append(path)
            print(status)
        num = num +1
num =2
#指定のクラスを リストで取得する。

情報抽出

for path in pathList:
    try:
        driver.get(str(path))
        WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)  

        soupに解析させる

        #カレントページを取得する
        html =driver.page_source.encode('utf-8')
        #読み込む情報を解析する。
        soup = BeautifulSoup(html, 'html.parser') 
        #注文番号取得                    
        orderNumber =  driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[1]/div[1]/div[2]/div/span')
        print(orderNumber.text)
        #商品名
        productName1 =  driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[1]')
        print(productName1.text)
        productName2 =  driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[2]')
        print(productName2.text)
        #status
        status = driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[2]/div/div/div/span')
        print(status.text)
        #数量
        quantity = driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[6]/span')
        print(quantity.text)
        #注文日
        store  =driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[1]/div[1]/div[1]/div/span')
        print(store.text)
        #住所　
        address       =driver.find_element(By.XPATH,'//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[1]/div/div')
        print(address.text)

        アイテム記入場所定義

        #注文番号取得
        inputWs.cell(row = num, column = yoko +1 ).value = str(orderNumber.text)
        #商品名
        inputWs.cell(row = num, column = yoko +2 ).value = str(productName1.text + productName2.text)
        #ステータス
        inputWs.cell(row = num, column = yoko +3 ).value = str(status.text)
        #数量
        inputWs.cell(row = num, column = yoko +4 ).value = str(quantity.text)
        #注文日
        inputWs.cell(row = num, column = yoko +5 ).value = str(store.text)
        #住所        
        inputWs.cell(row = num, column = yoko +7 ).value =  str(address.text)
        num = num + 1
    except:
        print("不明なエラー")
#ファイルディレクトリの取得
pwd = os.getcwd()
filePath = pwd + "\\" + outputFile
inputWb.save(filePath)
# メッセージボックス（情報） 
messagebox.showinfo('正常終了', '処理を終了します')
'''

"""